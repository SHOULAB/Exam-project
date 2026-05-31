# -*- coding: utf-8 -*-
"""
generate_test4.py
Generee Excel failu ar Budgetar sistemas testpiemeru kopam.
  - 1. lapa: Melnas kastes testpiemeru kopa (Black Box)
  - 2. lapa: Baltas kastes testpiemeru kopa (White Box)

ID kolonna: funkcionajas prasibas numurs (piem., 2.2.1).
"""

import sys
import subprocess

try:
    import openpyxl
except ImportError:
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'], check=True)
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT_PATH = Path(r'c:\Budgetar\Servera Faili\Budgetar\other\Sistemas_testesana.xlsx')

# ---------------------------------------------------------------------------
# Krasas un stili
# ---------------------------------------------------------------------------
HEADER_FILL  = PatternFill('solid', fgColor='1F3864')
TITLE_FILL   = PatternFill('solid', fgColor='DDEEFF')
ROW_ALT_FILL = PatternFill('solid', fgColor='DCE6F1')
ROW_FILL     = PatternFill('solid', fgColor='FFFFFF')

HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
DATA_FONT   = Font(name='Calibri', size=10)
ID_FONT     = Font(name='Calibri', size=10, bold=True)
TITLE_FONT  = Font(name='Calibri', size=13, bold=True, color='1F3864')

THIN  = Side(style='thin',   color='8EA9C1')
THICK = Side(style='medium', color='1F3864')

def make_border(l=False, r=False, t=False, b=False):
    return Border(
        left   = THICK if l else THIN,
        right  = THICK if r else THIN,
        top    = THICK if t else THIN,
        bottom = THICK if b else THIN,
    )

WRAP   = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

HEADERS = [
    'ID\n(prasibas\nnumurs)',
    'Izpildes nosacijumi',
    'Apraksts',
    'Izpildes soli',
    'Ievades dati',
    'Sagaidamais rezultats',
]
COL_WIDTHS = [9, 26, 26, 40, 26, 40]

HEADERS_WB = [
    'ID',
    'Izpildes nosacijumi',
    'Apraksts',
    'Izpildes soli',
    'Ievades dati',
    'Sagaidamais rezultats',
    'Prasibu\nidentifikators',
]
COL_WIDTHS_WB = [11, 24, 26, 38, 24, 38, 15]

# ---------------------------------------------------------------------------
# Melnas kastes testi
# ---------------------------------------------------------------------------
BLACK_BOX = [
    ('2.2.1',
     'Lietotajs nav registrets',
     'Veiksmiga registracija',
     '1) Atvert register.php\n2) Aizpildit lietotajvardu, e-pastu un paroli\n3) Atzimet privatuma politikas piekrisanu\n4) Nospiest "Registreties"',
     'Derig. lietotajvards (min.4), e-pasts, parole (min.8)',
     'Konts izveidots; lietotajs novirzits uz login.php'),

    ('2.2.1',
     'Nordita e-pasta adrese jau pastav datubaze',
     'Registracija ar jau esosu e-pastu',
     '1) Atvert register.php\n2) Ievadlt jau registretu e-pasta adresi\n3) Nospiest "Registreties"',
     'E-pasts, kas jau atrodas datubaze',
     'Kludas pazinojums "E-pasts jau ir registrets!"; konts netiek izveidots'),

    ('2.2.2',
     'Lietotajs ir registrets un aktives',
     'Veiksmiga pieteikšanās',
     '1) Atvert login.php\n2) Ievadlt pareizu e-pastu un paroli\n3) Nospiest "Ieiet"',
     'Pareizs e-pasts un parole',
     'Lietotajs autentificets un novirzits uz calendar.php'),

    ('2.2.2',
     'Lietotajs ir registrets',
     'Pieteikšanās ar nepareizu paroli',
     '1) Atvert login.php\n2) Ievadlt pareizu e-pastu un nepareizu paroli\n3) Nospiest "Ieiet"',
     'Pareizs e-pasts, nepareiza parole',
     'Kludas pazinojums "Nepareizs e-pasts vai parole!"; pieteikšanās nenotiek'),

    ('2.2.2',
     'Lietotaja konts ir deaktivets',
     'Pieteikšanās ar deaktivetu kontu',
     '1) Atvert login.php\n2) Ievadlt deaktiveta konta e-pastu un paroli\n3) Nospiest "Ieiet"',
     'Deaktiveta konta akreditacijas dati',
     'Kludas pazinojums; pieteikšanās nenotiek'),

    ('2.2.3',
     'Lietotajs ir registrets',
     'Paroles atjaunosanas e-pasta pieprasisana',
     '1) Atvert login.php\n2) Nospiest "Aizmirsu paroli"\n3) Ievadlt registreto e-pastu\n4) Nospiest "Nosutit"',
     'Registrets e-pasts',
     'Apstiprinajums paradzits; uz e-pastu nosutita atiestatisanas saite'),

    ('2.2.3',
     'Lietotajs sanemis atjaunosanas saiti',
     'Paroles atiestatisana ar derigu tokenu',
     '1) Atvert saiti no e-pasta\n2) Ievadlt jaunu paroli (min.8)\n3) Apstiprnat paroli\n4) Nospiest "Saglabat"',
     'Jauna parole, min. 8 rakstzimes',
     'Parole atjaunota; lietotajs novirzits uz login.php'),

    ('2.2.3',
     'Tokens ir beidzies (>1 stunda)',
     'Paroles atiestatisana ar beigusies tokenu',
     '1) Meginat atvert novecojusu atjaunosanas saiti',
     'Beidzies tokens URL',
     'Pazinojums "Saite ir nederiga vai deriguma laiks beidzies"'),

    ('2.2.4',
     'Lietotajs ir pieteicies',
     'Ienakumu ieraksta pievienosana',
     '1) Atvert calendar.php\n2) Nospiest "Pievienot"\n3) Izveleties tipu "Ienakumi"\n4) Aizpildit summu, aprakstu, datumu\n5) Nospiest "Saglabat"',
     'Summa:500, apraksts:"Alga", tips:ienakumi',
     'Ieraksts saglabats; redzams kalendara attiecigaja datuma'),

    ('2.2.4',
     'Lietotajs ir pieteicies',
     'Atkartota maksajuma pievienosana',
     '1) Pievienot jaunu ierakstu\n2) Ieslegt "Atkartojas maksajums"\n3) Izveleties atkartosanas dienas\n4) Saglabat',
     'Summa, apraksts, atkartosanas nedelas dienas',
     'Ieraksts saglabats ar atkartosanas atzimi; redzams vairakos datumos'),

    ('2.2.5',
     'Lietotajs ir pieteicies',
     'Izdevumu ieraksta pievienosana',
     '1) Atvert calendar.php\n2) Nospiest "Pievienot"\n3) Izveleties tipu "Izdevumi"\n4) Aizpildit laukus\n5) Nospiest "Saglabat"',
     'Summa:50, apraksts:"Partika", tips:izdevumi',
     'Ieraksts saglabats; redzams kalendara attiecigaja datuma'),

    ('2.2.5',
     'Summas lauks ir tukss',
     'Pievienosana ar tuksu summas lauku',
     '1) Nospiest "Pievienot"\n2) Atstit summas lauku tuksu\n3) Nospiest "Saglabat"',
     'Tukss summas lauks',
     'Validacijas kludas pazinojums; ieraksts netiek saglabats'),

    ('2.2.6',
     'Lietotajam ir vismaz viens esoss ieraksts',
     'Ieraksta redigesana',
     '1) Atvert calendar.php, noklikskinat uz dienas\n2) Nospiest redigesanas pogu\n3) Mainit summu un aprakstu\n4) Nospiest "Saglabat"',
     'Jauna summa:600, jauns apraksts:"Alga + piemaksa"',
     'Ieraksts atjauninats; izmaiñas redzamas kalendara'),

    ('2.2.7',
     'Lietotajam ir vismaz viens esoss ieraksts',
     'Ieraksta dzesana',
     '1) Atvert dienas skatu kalendara\n2) Nospiest dzesanas pogu pie ieraksta\n3) Apstiprnat',
     'Nav ievades datu',
     'Ieraksts dzests; vairs nav redzams kalendara'),

    ('2.2.8',
     'Lietotajam ir ieraksti vismaz viena menesi',
     'Statistikas parskatu skatiešanā',
     '1) Atvert parskati.php\n2) Parbaudit ienakumu/izdevumu grafikas par pedejiem 12 menesiem',
     'Nav ievades datu',
     'Grafiki attelo pareizus datus; katrs menesis rada attiecigajas summas'),

    ('2.2.9',
     'Lietotajam ir ieraksti ar dazadam kategorijam',
     'Filtresana pec kategorijas',
     '1) Atvert transakciju sarakstu\n2) Izveleties kategoriju "Partika" filtra\n3) Piemerot filtru',
     'Kategorija: "Partika"',
     'Tiek attalotas tikai kategorijas "Partika" ieraksti'),

    ('2.2.9',
     'Lietotajam ir ieraksti ar aprakstiem',
     'Meklesana pec atslēgvārda',
     '1) Meklesanas lauka ievadlt atslēgvārdu\n2) Apstiprnat meklesanu',
     'Atslēgvārds: "alga"',
     'Redzami tikai ieraksti, kuru apraksta ir vards "alga"'),

    ('2.2.10',
     'Ir vairaki ienakumu un izdevumu ieraksti',
     'Informativo datu apreksins (bilance, ienakumi, izdevumi)',
     '1) Pievienot ienakumus EUR 500 un EUR 200\n2) Pievienot izdevumu EUR 300\n3) Atvert informacijas paneli',
     'Ienakumi:EUR 700, Izdevumi:EUR 300',
     'Paneli rada: Ienakumi EUR 700, Izdevumi EUR 300, Bilance EUR 400'),

    ('2.2.11',
     'Lietotajs ir pieteicies',
     'Budzeta plana izveide',
     '1) Atvert budget.php\n2) Nospiest "Pievienot budzetu"\n3) Aizpildit nosaukumu, summu, periodu\n4) Saglabat',
     'Nosaukums:"Maija budzets", summa:EUR 500',
     'Budzeta plans izveidots un redzams saraksta'),

    ('2.2.11',
     'Lietotajam ir esoss budzeta plans',
     'Budzeta plana redigešanā',
     '1) Nospiest redigesanas pogu\n2) Mainit summu uz EUR 600\n3) Saglabat',
     'Jauna summa: EUR 600',
     'Budzets atjauninats ar jauno summu'),

    ('2.2.11',
     'Lietotajam ir esoss budzeta plans',
     'Budzeta plana dzesana',
     '1) Nospiest dzesanas pogu pie budzeta\n2) Apstiprnat dzesanu',
     'Nav ievades datu',
     'Budzets dzests; vairs nav redzams saraksta'),

    ('2.2.12',
     'Aktives budzets EUR 200 menesim',
     'Bridinajuma pazinojums (>80% no budzeta)',
     '1) Pievienot izdevumus EUR 170 (85% no EUR 200)\n2) Atsvaidzinat lapu',
     'Izdevumi: EUR 170 (>80% no EUR 200)',
     'Bridinajuma logs par deficitu; norādīts parsniegtais limits'),

    ('2.2.13',
     'Lietotajs ir registrets',
     '"Atcereties mani" automatiska pieteikšanās',
     '1) Pieteikties ar atzmetu "Atcereties mani"\n2) Aizvrt parlukprogrammu\n3) Atkartoti atvert vietni',
     'Pareizs e-pasts, parole, atzmeta opcija',
     'Lietotajs automatiski pieteikts bez atkartots paroles ievadisanas'),

    ('2.2.14',
     'Ir sagatavots derigs CSV fails',
     'Transakciju imports no CSV faila',
     '1) Atvert settings.php\n2) Sadata "Imports" augšupielādēt derigu CSV\n3) Izveleties "Transakcijas"\n4) Nospiest "Importet"',
     'Derigs CSV ar kolonnu nosaukumiem un datiem',
     'Apstiprinajums ar importeto ierakstu skaitu; dati redzami sistema'),

    ('2.2.14',
     'Lietotajs ir pieteicies',
     'Imports ar nepareizu faila tipu',
     '1) Atvert settings.php\n2) Meginat augšupielādēt .xlsx failu\n3) Nospiest "Importet"',
     'Fails: .xlsx (nepareizs formats)',
     'Kludas pazinojums "Nepareizs faila formats"; dati netiek importeti'),

    ('2.2.15',
     'Lietotajam ir ieraksti datubaze',
     'Transakciju eksports uz CSV',
     '1) Atvert settings.php\n2) Izveleties "Transakcijas"\n3) Nospiest "Eksportet"',
     'Eksporta tips: Transakcijas',
     'CSV fails lejupielādēts; satur visus ierakstus ar pareiziem laukiem'),

    ('2.2.15',
     'Lietotajam ir budzeti datubaze',
     'Budzetu eksports uz CSV',
     '1) Atvert settings.php\n2) Izveleties "Budzeti"\n3) Nospiest "Eksportet"',
     'Eksporta tips: Budzeti',
     'CSV fails lejupielādēts; satur visus budzetus'),

    ('2.2.16',
     'Lietotajs nav pieteicies',
     'Piekļuve aizsargātai lapai bez autorizacijas',
     '1) Izrakstities\n2) Manuali ievadlt parluka URL: calendar.php',
     'Nav ievades datu',
     'Lietotajs novirzits uz login.php; saturs nav redzams'),

    ('2.2.17',
     'Lietotajs ir pieteicies',
     'Lietotajvarda maina',
     '1) Atvert settings.php\n2) Ievadlt jaunu lietotajvardu\n3) Ievadlt pašreizejo paroli\n4) Saglabat',
     'Jauns unikalss lietotajvards, pareiza esosa parole',
     'Lietotajvards atjauninats; apstiprinajums radzams'),

    ('2.2.17',
     'Lietotajs ir pieteicies',
     'Paroles maina',
     '1) Atvert settings.php\n2) Ievadlt esoso paroli\n3) Ievadlt jauno paroli divreiz\n4) Saglabat',
     'Pareiza esosa parole, jauna parole (min.8)',
     'Parole veiksmigi atjaunota; apstiprinajums radzams'),

    ('2.2.17',
     'Lietotajs ir pieteicies',
     'Paroles maina ar nepareizu esoso paroli',
     '1) Atvert settings.php\n2) Ievadlt nepareizu esoso paroli\n3) Saglabat',
     'Nepareiza esosa parole',
     'Kludas pazinojums; parole netiek mainita'),

    ('2.2.18',
     'Lietotajs ir pieteicies',
     'Veiksmiga izrakstišanās',
     '1) Nospiest "Iziet" sanjosla\n2) Apstiprnat dialoglodziña',
     'Nav ievades datu',
     'Sesija dzesta; lietotajs novirzits uz index.php'),

    ('2.2.19',
     'Lietotajs ir pieteicies',
     'Valutas maina no EUR uz USD',
     '1) Atvert settings.php\n2) Izveleties "USD"\n3) Saglabat',
     'Valuta: USD',
     'Visas summas sistema attalotas ar USD simbolu'),

    ('2.2.20',
     'Lietotajs ir pieteicies',
     'Valodas maina no latviesu uz anglu',
     '1) Atvert settings.php\n2) Izveleties "English"\n3) Saglabat',
     'Valoda: English',
     'Sistemas saskarne parsledzas uz anglu valodu'),

    ('2.2.21',
     'Lietotajs nav registrets',
     'Registracija bez privatuma politikas piekrisanas',
     '1) Atvert register.php\n2) Aizpildit laukus\n3) NEATZIMET privatuma politiku\n4) Nospiest "Registreties"',
     'Derigi lauki, privatuma politika neatzmeta',
     'Registracija netiek pabeigta; kludas pazinojums par nepieciesamo piekrisanu'),

    ('2.2.22',
     'Lietotajs ir pieteicies',
     'Parslegt uz tumso krasu shemu',
     '1) Atvert settings.php\n2) Izveleties "Tumsa"\n3) Saglabat',
     'Krasu shema: tumsa',
     'Sistemas fons parsledzas uz tumso shemu; iestatijums saglabats'),

    ('2.2.23',
     'Ir ieraksti dazados menesus',
     'Navigacija starp menesiem un dienas skats',
     '1) Atvert calendar.php\n2) Nospiest nakama menesa bultiñu\n3) Nospiest iepriekseja menesa bultiñu\n4) Noklikskinat uz dienas ar ierakstiem',
     'Nav ievades datu',
     'Paredzamais menesis attālojas; dienas skats attalo visus ierakstus'),

    ('2.2.24',
     'Mobila ierice ar HTTPS pieejama sistema',
     'Lietotnes instalacija ka PWA',
     '1) Atvert vietni mobila parluka\n2) Apstiprnat "Pievienot sakuma ekranam"',
     'Nav ievades datu',
     'Lietotne pievienota sakuma ekranam; palais standalone rezima bez parluka UI'),

    ('2.2.25',
     'Ir finansu dati; lietotajs ir pieteicies',
     'Konta atiestatisana – visu finansu datu dzesana',
     '1) Atvert settings.php\n2) Ievadlt paroli\n3) Nospiest "Atiestatit"\n4) Apstiprnat',
     'Pareiza parole',
     'Visi transakciju un budzetu ieraksti dzesti; konts un iestatijumi saglabati'),

    ('2.2.25',
     'Lietotajs ir pieteicies',
     'Konta atiestatisana ar nepareizu paroli',
     '1) Atvert settings.php\n2) Ievadlt nepareizu paroli\n3) Nospiest "Atiestatit"',
     'Nepareiza parole',
     'Kludas pazinojums; dati netiek dzesti'),

    ('2.2.26',
     'Pieteicies ka administrators vai moderators',
     'Lietotaja datu redigesana admin paneli',
     '1) Atvert admin paneli\n2) Nospiest "Rediget"\n3) Mainit lietotajvardu\n4) Saglabat',
     'Jauns unikalss lietotajvards',
     'Dati atjauninati; izmaiñas redzamas paneli'),

    ('2.2.26',
     'Pieteicies ka moderators',
     'Moderators nevar mainit lietotaja lomu',
     '1) Atvert admin paneli\n2) Meginat mainit lomu uz "moderator"',
     'Lomas lauks',
     'Lomas lauks nav redizejams moderatoram; lomu var mainit tikai administrators'),

    ('2.2.27',
     'Merkis ir aktivs "user"; pieteicies ka admin/mod',
     'Lietotaja konta deaktivasana',
     '1) Atvert admin paneli\n2) Nospiest "Deaktivet"\n3) Apstiprnat',
     'Lietotaja ID ar lomu "user"',
     'Konts deaktivets (is_active=0); lietotajs nevar pieteikties; dati saglabati'),

    ('2.2.27',
     'Pieteicies ka moderators; merkis ir moderators',
     'Moderators nevar blokot citu moderatoru',
     '1) Atvert admin paneli\n2) Meginat deaktivet moderatoru',
     'Moderatora ID',
     'Darbiba nav atlauta; kludas pazinojums; konts paliek aktivs'),

    ('2.2.28',
     'Merkis ir deaktivets konts; pieteicies ka administrators',
     'Deaktiveta lietotaja konta dzesana',
     '1) Atvert admin paneli\n2) Atrast neaktivu kontu\n3) Nospiest "Dzest"\n4) Apstiprnat',
     'Lietotaja ID ar is_active=0',
     'Visi dati (transakcijas, budzeti, iestatijumi, konts) neatgriezeniski dzesti'),

    ('2.2.28',
     'Lietotajs ir pieteicies',
     'Konta pašdzesana',
     '1) Atvert settings.php\n2) Ievadlt paroli\n3) Nospiest "Dzest kontu"\n4) Apstiprnat',
     'Pareiza parole',
     'Visi dati un konts dzesti; lietotajs novirzits uz login.php'),
]

# ---------------------------------------------------------------------------
# Baltas kastes testi
# ---------------------------------------------------------------------------
WHITE_BOX = [
    # --- Autentifikacija un drošiba ---
    ('TP.WB.01',
     'Datubaze ir pieejama; ir registrets lietotajs',
     'Paroles hešošanas parbaude (password_hash)',
     '1) Registreties ar jaunu kontu, parole "test1234"\n2) Atvert phpMyAdmin\n3) Parbaudit lauku "password" BU_users tabula jaunajam kontam',
     'Parole: "test1234"',
     'Datubaze lauka "password" ir hešota vertiba (sakas ar $2y$), nevis atklats teksts "test1234"',
     '2.2.1'),

    ('TP.WB.02',
     'Pieteikšanās forma ir pieejama',
     'SQL injekcijas aizsardziba ar sagatavotas komandas (prepared statement)',
     '1) Atvert login.php\n2) Lauka "e-pasts" ievadlt: \' OR \'1\'=\'1\n3) Lauka "parole" ievadlt jebko\n4) Nospiest "Ieiet"',
     "E-pasts: ' OR '1'='1, parole: jebkas",
     'Sistema nepieteic; parāda kludas pazinojumu. Koda līmenī mysqliStmt prepare tiek izmantots — ievade tiek apstrādāta kā parametrs, nevis SQL koda daļa',
     '2.2.2'),

    ('TP.WB.03',
     'Lietotajs ir pieteicies, ta sesija ir aktiva',
     'Sesijas parbaude aizsargatas lapas (auth_check.php)',
     '1) Piesakties kā derīgs lietotājs\n2) Kopēt sesijas sīkfailu\n3) Izrakstities (sesija dzēsta serverī)\n4) Manuāli ievadīt parlūkā: /user/php/calendar.php',
     'Derīga sesijas sīkfaila vērtība, bet serverī sesija dzēsta',
     'auth_check.php pārbauda sesijas mainīgo $_SESSION["user_id"]; lietotājs tiek novirzīts uz login.php — saturs nav redzams',
     '2.2.16'),

    ('TP.WB.04',
     'Lietotājs nav autentificēts',
     'Tieša piekļuve aizsargātai lapai bez sesijas',
     '1) Pārliecināties, ka nav aktīvas sesijas (inkognito logs)\n2) Pārlūkā ievadīt: /user/php/budget.php',
     'Nav ievades datu (sesija nepastāv)',
     'auth_check.php konstatē, ka $_SESSION["user_id"] nav iestatīts; lietotājs tiek novirzīts uz login.php',
     '2.2.16'),

    ('TP.WB.05',
     'Lietotājs piesakās ar atzīmētu "Atcerēties mani"',
     '"Atcerēties mani" — tokena ģenerēšanas loģika (unikāls, drošs)',
     '1) Atvert login.php\n2) Ievadīt pareizus akreditācijas datus\n3) Atzīmēt "Atcerēties mani"\n4) Nospiest "Ieiet"\n5) Atvērt phpMyAdmin → tabulu BU_remember_tokens',
     'Pareizs e-pasts un parole, atzīmēta opcija',
     'BU_remember_tokens tabulā parādās jauns ieraksts ar user_id, nejaušu tokenu (min. 32 rakstzīmes), expires_at (pašreizējais laiks + 30 dienas) un created_at',
     '2.2.13'),

    ('TP.WB.06',
     'BU_remember_tokens tabulā ir derīgs tokens; lietotājs atver vietni',
     '"Atcerēties mani" — tokena validācija pēc pārlūka restartēšanas',
     '1) Piesakties ar "Atcerēties mani"\n2) Aizvērt pārlūkprogrammu pilnībā\n3) Atkārtoti atvērt vietni',
     'Derīgs tokens sīkfailā (expires_at nākotnē)',
     'login.php pārbauda sīkfaila tokenu pret BU_remember_tokens; token derīgs → lietotājs tiek autentificēts automātiski bez paroles ievades',
     '2.2.13'),

    ('TP.WB.07',
     'Lietotājs ar "Atcerēties mani" sesiju; veic izrakstīšanos',
     '"Atcerēties mani" tokena dzēšana pēc izrakstīšanās',
     '1) Piesakties ar "Atcerēties mani"\n2) Nospiest "Iziet"\n3) Pārbaudīt BU_remember_tokens tabulu phpMyAdmin\n4) Pārbaudīt pārlūkprogrammas sīkfailus',
     'Nav papildu ievades datu',
     'Atbilstošais ieraksts BU_remember_tokens tabulā ir dzēsts; pārlūkprogrammas sīkfails "remember_token" ir noņemts (expires_at pagātne)',
     '2.2.18'),

    ('TP.WB.08',
     'Pieteiktas sesija ir aktīva',
     'XSS (Cross-Site Scripting) novēršana — htmlspecialchars() izvadē',
     '1) Pievienot transakciju ar aprakstu: <script>alert("XSS")</script>\n2) Saglabāt\n3) Atvērt kalendāru un apskatīt ierakstu',
     'Apraksts: <script>alert("XSS")</script>',
     'Pārlūkprogramma neparāda uznirstošo logu; apraksts tiek attēlots kā teksts &lt;script&gt;alert("XSS")&lt;/script&gt; — htmlspecialchars() ir piemērots izvadē',
     '2.2.16'),

    # --- Datu šifrešana (AES-256-CBC) ---
    ('TP.WB.09',
     'Lietotājs ir pieteicies; datubāze pieejama',
     'AES-256-CBC šifrēšana — transakcijas summas glabāšana',
     '1) Pievienot transakciju ar summu 123.45\n2) Atvērt phpMyAdmin → BU_transactions tabula\n3) Apskatīt lauku "amount" jaunākajam ierakstam',
     'Summa: 123.45',
     'Lauka "amount" vērtība datubāzē ir base64 kodēta šifrēta virkne (nav atpazīstams skaitlis 123.45) — AES-256-CBC šifrēšana ir piemērota',
     '2.2.4'),

    ('TP.WB.10',
     'BU_transactions tabulā ir šifrēti ieraksti',
     'AES-256-CBC atšifrēšana — korekta summas attēlošana',
     '1) Pievienot transakciju ar summu 999.99\n2) Atjaunināt kalendāra lapu\n3) Atvērt dienas skatu',
     'Summa pirms šifrēšanas: 999.99',
     'Kalendārs attēlo skaitli 999.99 — atšifrēšanas loģika calendar.php darbojas korekti un atgriež oriģinālo vērtību',
     '2.2.23'),

    ('TP.WB.11',
     'Ir vairāki transakciju ieraksti datubāzē',
     'Unikāls IV katrai šifrēšanas operācijai',
     '1) Pievienot divas transakcijas ar vienādu summu (100.00) un aprakstu\n2) Atvērt phpMyAdmin → BU_transactions\n3) Salīdzināt lauku "amount" abiem ierakstiem',
     'Divas transakcijas ar summu 100.00',
     'Lauka "amount" šifrētās vērtības abiem ierakstiem atšķiras — katrā šifrēšanas reizē tiek ģenerēts jauns IV, novēršot identificējamas rakstveida',
     '2.2.16'),

    # --- Finansu apreksini ---
    ('TP.WB.12',
     'Ir vairāki ienākumu un izdevumu ieraksti tekošajā mēnesī',
     'Bilances aprēķina pareizība (ienākumi − izdevumi)',
     '1) Pievienot ienākumus: EUR 500, EUR 200\n2) Pievienot izdevumus: EUR 150, EUR 100\n3) Atvērt informācijas paneli',
     'Ienākumi: EUR 700, Izdevumi: EUR 250',
     'Informācijas panelis rāda: Ienākumi EUR 700.00, Izdevumi EUR 250.00, Bilance EUR 450.00 — aprēķins 700 − 250 = 450 ir pareizs',
     '2.2.10'),

    ('TP.WB.13',
     'Ir ieraksti vairākos mēnešos',
     'Mēneša kopējo ienākumu un izdevumu aprēķins statistikā',
     '1) Pievienot janvārī ienākumus EUR 1200 un izdevumus EUR 800\n2) Pievienot februārī ienākumus EUR 900 un izdevumus EUR 600\n3) Atvērt parskati.php',
     'Janvāris: ienākumi EUR 1200, izdevumi EUR 800; februāris: ienākumi EUR 900, izdevumi EUR 600',
     'Statistikas grafiki rāda korektas summas katram mēnesim — SQL vaicājums ar GROUP BY MONTH() apkopo datus pareizi',
     '2.2.8'),

    ('TP.WB.14',
     'Aktīvs budžets EUR 200 mēnesim; pieteikts lietotājs',
     'Budžeta izlietojuma procentu aprēķins (80% brīdinājuma slieksnis)',
     '1) Izveidot budžetu EUR 200 tekošajam mēnesim\n2) Pievienot izdevumus EUR 160 (80% no 200)\n3) Atvērt budget.php',
     'Budžets: EUR 200; izdevumi: EUR 160 (tieši 80%)',
     'Sistēma attēlo brīdinājuma paziņojumu (160/200 = 80% ≥ 80%); brīdinājuma slieksnis tiek aprēķināts ar: (izdevumi / budžets_summa) * 100 ≥ 80',
     '2.2.12'),

    # --- Piekluves kontrole ---
    ('TP.WB.15',
     'Divi lietotāji ar datiem (Lietotājs A un Lietotājs B)',
     'Lietotāja datu izolācija — cita lietotāja datu aizsardzība',
     '1) Piesakties kā Lietotājs A\n2) Iegūt Lietotāja B pievienotas transakcijas ID\n3) Manuāli mainīt URL: calendar.php?id=[B_transakcijas_ID]\n4) Novērot atbildi',
     'URL ar cita lietotāja transakcijas ID',
     'Sistēma pārbauda, ka ieraksta user_id atbilst sesijas user_id; pieprasījums tiek noraidīts vai atgriež tukšu rezultātu — cita lietotāja dati nav redzami',
     '2.2.16'),

    ('TP.WB.16',
     'Pieteicies kā administrators; mērķis ir cits administrators',
     'Administrators nevar rediģēt cita administratora kontu',
     '1) Piesakties kā Administrators A\n2) Atvērt admin paneli\n3) Mēģināt rediģēt Administratora B kontu',
     'Mērķa konta loma: "admin"',
     'Sistēma pārbauda mērķa lietotāja lomu — ja loma ir "admin", darbība tiek noraidīta; tiek parādīts kļūdas paziņojums',
     '2.2.26'),

    ('TP.WB.17',
     'Pieteicies kā moderators; mērķis ir cits moderators',
     'Moderators nevar bloķēt/dzēst citu moderatoru vai administratoru',
     '1) Piesakties kā Moderators\n2) Atvērt admin paneli\n3) Mēģināt deaktivēt Moderatora kontu',
     'Mērķa konta loma: "moderator"',
     'Koda līmenī tiek pārbaudīts: ja pašreizējais lietotājs ir "moderator" un mērķa loma nav "user", darbība tiek bloķēta; tiek parādīts kļūdas paziņojums',
     '2.2.27'),

    ('TP.WB.18',
     'Deaktivēts lietotāja konts (is_active=0)',
     'Deaktivēta konta piekļuves bloķēšana pieteikšanās laikā',
     '1) Deaktivēt lietotāja kontu (is_active=0)\n2) Mēģināt piesakties ar šī konta akreditācijas datiem',
     'Pareizs e-pasts un parole no deaktivēta konta',
     'login.php vaicājums pārbauda is_active=1; deaktivētam kontam tiek atgriezts kļūdas paziņojums — pieteikšanās nenotiek pat ar pareiziem akreditācijas datiem',
     '2.2.2'),

    # --- Datu validacija (ievades parbaude) ---
    ('TP.WB.19',
     'Lietotājs ir pieteicies; transakcijas forma atvērta',
     'Negatīvas summas noraidīšana (server-side validācija)',
     '1) Atvērt transakcijas pievienošanas formu\n2) Ar pārlūka DevTools izslēgt front-end validāciju\n3) Summas laukā ievadīt -50\n4) Nospiest "Saglabāt"',
     'Summa: -50',
     'Server-side kods (budget.php vai calendar.php) pārbauda, vai summa > 0; negatīva vērtība tiek noraidīta un tiek atgriezts kļūdas paziņojums — datubāzē ieraksts netiek saglabāts',
     '2.2.4'),

    ('TP.WB.20',
     'Lietotājs ir pieteicies; transakcijas forma atvērta',
     'Obligāto lauku server-side validācija',
     '1) Ar pārlūka DevTools izslēgt front-end validāciju\n2) Atstāt summas lauku tukšu\n3) Nospiest "Saglabāt"',
     'Tukšs summas lauks (pēc front-end apvedceļa)',
     'PHP kods pārbauda, vai obligātie lauki (summa, datums, tips) ir aizpildīti; tukša vērtība tiek noraidīta — datubāzē ieraksts netiek saglabāts, tiek atgriezta kļūda',
     '2.2.5'),

    ('TP.WB.21',
     'Reģistrācijas forma ir pieejama',
     'Paroles minimālā garuma server-side validācija',
     '1) Ar DevTools izslēgt front-end validāciju\n2) Aizpildīt reģistrācijas formu ar paroli "abc" (3 rakstzīmes)\n3) Nospiest "Reģistrēties"',
     'Parole: "abc" (mazāka par 8 rakstzīmēm)',
     'PHP kods pārbauda strlen($password) >= 8; parole ar 3 rakstzīmēm tiek noraidīta, tiek parādīts kļūdas paziņojums — konts netiek izveidots',
     '2.2.1'),

    # --- CSV imports/eksports ---
    ('TP.WB.22',
     'Lietotājs ir pieteicies; settings.php ir atvērts',
     'CSV faila paplašinājuma validācija importā',
     '1) Mēģināt augšupielādēt failu ar paplašinājumu .xlsx\n2) Nospiest "Importēt"',
     'Fails: test.xlsx',
     'PHP pārbauda faila paplašinājumu (pathinfo($file, PATHINFO_EXTENSION)); .xlsx tiek noraidīts — tiek parādīts kļūdas paziņojums "Nepareizs faila formāts"',
     '2.2.14'),

    ('TP.WB.23',
     'Lietotājs ir pieteicies; settings.php ir atvērts',
     'CSV faila MIME tipa validācija importā',
     '1) Pārdēvēt .exe failu par test.csv\n2) Mēģināt augšupielādēt\n3) Nospiest "Importēt"',
     'Fails: test.csv (patiesais tips: application/octet-stream)',
     'PHP pārbauda $_FILES["file"]["type"] vai mime_content_type(); nepareizs MIME tips tiek noraidīts — ieraksti netiek importēti',
     '2.2.14'),

    ('TP.WB.24',
     'CSV failā ir nepareiza kolonnu struktūra',
     'CSV kolonnu struktūras validācija importā',
     '1) Sagatavot CSV failu ar nepareiziem kolonnu nosaukumiem\n2) Augšupielādēt\n3) Nospiest "Importēt"',
     'CSV fails ar nepareiziem virsrakstiem (piem., "date" vietā "datum")',
     'PHP pārbauda sagaidāmos kolonnu nosaukumus; neatbilstoša struktūra tiek noraidīta — tiek parādīts kļūdas paziņojums, dati netiek importēti',
     '2.2.14'),

    ('TP.WB.25',
     'Lietotājam ir transakciju ieraksti datubāzē',
     'CSV eksportā — šifrēto lauku atšifrēšana pirms rakstīšanas',
     '1) Pievienot transakciju ar summu 250.00 un aprakstu "Alga"\n2) Atvērt settings.php\n3) Nospiest "Eksportēt" (transakcijas)\n4) Atvērt lejupielādēto CSV failu',
     'Transakcijas summa: 250.00, apraksts: "Alga"',
     'CSV failā ir skaitlis 250.00 un teksts "Alga" — nevis šifrēta virkne; datu eksporta PHP kods atšifrē "amount" un "description" laukus pirms rakstīšanas CSV',
     '2.2.15'),

    ('TP.WB.26',
     'Lietotājam ir transakciju ieraksti datubāzē',
     'CSV eksportā — UTF-8 kodējuma pārbaude (latviešu rakstzīmes)',
     '1) Pievienot transakciju ar aprakstu "Ēdiens, Sērkociņi"\n2) Eksportēt transakcijas kā CSV\n3) Atvērt CSV failā ar teksta redaktoru (pārbaudīt raw bytes faila sākumā)',
     'Apraksts ar latviešu diakritiskajām zīmēm',
     'CSV faila pirmās trīs baiti ir EF BB BF (UTF-8 BOM); latviešu rakstzīmes (ā, ē, ī, u.c.) tiek attēlotas pareizi — PHP ieraksta chr(0xEF).chr(0xBB).chr(0xBF) pirms satura',
     '2.2.15'),

    # --- Atkārtotās transakcijas ---
    ('TP.WB.27',
     'Atkārtota transakcija ar norādītām nedēļas dienām',
     'Atkārtotās transakcijas loģika — parādīšanās pareizās dienās',
     '1) Pievienot transakciju ar "Atkartojas" ieslēgtu, atzīmēt: pirmdiena un ceturtdiena\n2) Atvērt nākamo nedēļu kalendārā\n3) Pārbaudīt, kurās dienās ieraksts ir redzams',
     'Atkārtošanās dienas: pirmdiena (1), ceturtdiena (4)',
     'Kalendāra PHP kods atšifrē "recurring_days" lauku (JSON vai bitmask) un attēlo ierakstu tikai pirmdienās un ceturtdienās — citās dienās ieraksts nav redzams',
     '2.2.4'),

    ('TP.WB.28',
     'Atkārtota transakcija ar iestatītu apstāšanās datumu',
     'Atkārtotās transakcijas apstāšanās datuma ievērošana',
     '1) Pievienot atkārtotu transakciju ar recurring_stop_date = šī mēneša beigas\n2) Pārvietoties uz nākamo mēnesi kalendārā',
     'recurring_stop_date: pašreizējā mēneša pēdējā diena',
     'Nākamajā mēnesī ieraksts nav redzams — PHP kods pārbauda: ja current_date > recurring_stop_date, transakcija netiek iekļauta vaicājuma rezultātos',
     '2.2.4'),

    # --- Datubāzes integritāte ---
    ('TP.WB.29',
     'Administrators ir pieteicies; mērķa konts ir deaktivēts',
     'ON DELETE CASCADE — kaskādes dzēšana lietotāja dzēšanā',
     '1) Pārliecināties, ka mērķa lietotājam ir transakcijas, budžeti, iestatījumi, tokeni\n2) Admina panelī dzēst šo lietotāju\n3) Pārbaudīt phpMyAdmin: BU_transactions, BU_budgets, BU_user_settings, BU_remember_tokens, BU_password_resets',
     'Mērķa lietotāja ID ar saistītajiem ierakstiem',
     'Visas saistītās tabulas ir notīrītas: BU_transactions, BU_budgets, BU_user_settings tiek dzēstas ar PHP kodu; BU_remember_tokens un BU_password_resets ar ON DELETE CASCADE — neviens bāreņa ieraksts nepaliek',
     '2.2.28'),

    ('TP.WB.30',
     'Lietotājs ir saņēmis paroles atiestatīšanas saiti',
     'Paroles atiestatīšanas tokena vienreizlietošana',
     '1) Nosūtīt paroles atiestatīšanas pieprasījumu\n2) Atvērt saiti un ievadīt jaunu paroli\n3) Nospiest "Saglabāt"\n4) Mēģināt izmantot to pašu saiti vēlreiz',
     'Tas pats atiestatīšanas URL (tokens)',
     'Pēc veiksmīgas paroles maiņas kolonna "used" BU_password_resets tiek iestatīta uz 1; otro reizi izmantojot saiti, sistēma konstatē used=1 un parāda "Saite ir nederīga vai derīguma laiks beidzies"',
     '2.2.3'),

    ('TP.WB.31',
     'Paroles atiestatīšanas tokens ir vecāks par 1 stundu',
     'Paroles atiestatīšanas tokena derīguma termiņa pārbaude',
     '1) Datubāzē manuāli iestatīt BU_password_resets.expires_at uz pagātnes vērtību (piem., NOW() - 2h)\n2) Mēģināt izmantot atbilstošo atiestatīšanas saiti',
     'Beidzies tokens URL (expires_at < NOW())',
     'PHP kods vaicā: WHERE token = ? AND expires_at > NOW() AND used = 0; beidzies tokens neatbilst nosacījumam — tiek parādīts paziņojums "Saite ir nederīga vai derīguma laiks beidzies"',
     '2.2.3'),
]



# ---------------------------------------------------------------------------
# Excel rakstisana
# ---------------------------------------------------------------------------
def style_cell(cell, font=None, fill=None, alignment=None, border=None):
    if font:      cell.font = font
    if fill:      cell.fill = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border = border


def write_sheet(ws, title, headers, col_widths, rows):
    ncols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(row=1, column=1, value=title)
    style_cell(tc, font=TITLE_FONT, fill=TITLE_FILL,
               alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 28

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        style_cell(cell, font=HEADER_FONT, fill=HEADER_FILL,
                   alignment=CENTER,
                   border=make_border(l=(ci==1), r=(ci==ncols), t=True, b=True))
    ws.row_dimensions[2].height = 44

    nrows = len(rows)
    for ri, row in enumerate(rows, 3):
        fill = ROW_ALT_FILL if ri % 2 == 0 else ROW_FILL
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=str(val))
            style_cell(cell,
                       font=ID_FONT if ci == 1 else DATA_FONT,
                       fill=fill,
                       alignment=CENTER if ci == 1 else WRAP,
                       border=make_border(l=(ci==1), r=(ci==ncols),
                                          b=(ri == nrows + 2)))
        ws.row_dimensions[ri].height = 85

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A3'


def main():
    wb = openpyxl.Workbook()

    ws_bb = wb.active
    ws_bb.title = 'Melnas kastes testi'
    write_sheet(ws_bb,
                'Meln\u0101s kastes testpiem\u0113ru kopa',
                HEADERS, COL_WIDTHS, BLACK_BOX)

    ws_wb = wb.create_sheet('Baltas kastes testi')
    write_sheet(ws_wb,
                'Balt\u0101s kastes testpiem\u0113ru kopa',
                HEADERS_WB, COL_WIDTHS_WB, WHITE_BOX)

    wb.save(str(OUTPUT_PATH))
    print('Saglabats: ' + str(OUTPUT_PATH))
    print('  Melnas kastes testi:  ' + str(len(BLACK_BOX)) + ' rindas')
    print('  Baltas kastes testi:  ' + str(len(WHITE_BOX)) + ' rindas')


if __name__ == '__main__':
    main()
